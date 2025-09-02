import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import platform

# Core logic
def get_file_details(folder_path):
    file_data = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                stats = os.stat(file_path)
                size = stats.st_size
                mod_time = datetime.fromtimestamp(stats.st_mtime)
                create_time = datetime.fromtimestamp(stats.st_ctime)
                relative_folder = os.path.relpath(root, folder_path)

                file_data.append({
                    "Folder": relative_folder,
                    "File Name": file,
                    "Full Path": file_path,
                    "Size (Bytes)": size,
                    "Modified Date": mod_time.strftime('%Y-%m-%d %H:%M:%S'),
                    "Created Date": create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    "Extension": os.path.splitext(file)[1],
                    "Is Hidden": file.startswith('.'),
                    "Is Readable": os.access(file_path, os.R_OK),
                    "Is Writable": os.access(file_path, os.W_OK),
                })

            except Exception as e:
                print(f"Error processing {file_path}: {e}")

    return file_data


def export_to_excel(data, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Folder File Details"

    if not data:
        return False

    headers = list(data[0].keys())
    ws.append(headers)

    for item in data:
        ws.append([item[col] for col in headers])

    for col_num, col_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell)) for cell in [col_title] + [row[col_title] for row in data])
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(save_path)
    return True


# GUI logic
def browse_folder():
    folder_selected = filedialog.askdirectory(title="Select Folder to Scan")
    if not folder_selected:
        return

    status_var.set("Scanning folder...")
    window.update_idletasks()

    file_details = get_file_details(folder_selected)

    if not file_details:
        messagebox.showinfo("No Files", "No files found in the selected folder.")
        status_var.set("No files found.")
        return

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Excel File As"
    )

    if save_path:
        success = export_to_excel(file_details, save_path)
        if success:
            messagebox.showinfo("Success", f"Excel file saved successfully:\n{save_path}")
            status_var.set("Excel saved successfully.")
        else:
            messagebox.showerror("Error", "Failed to save Excel file.")
            status_var.set("Failed to save Excel.")
    else:
        status_var.set("Save cancelled.")


# ------------------------
# GUI Setup (Styled)
# ------------------------
window = tk.Tk()
window.title("Folder File Exporter")
window.geometry("500x300")
window.resizable(False, False)

# Use native icons if available
if platform.system() == "Windows":
    window.iconbitmap(default='')  # Replace with path to your .ico if needed

# Fonts and Styles
HEADER_FONT = ("Segoe UI", 16, "bold")
LABEL_FONT = ("Segoe UI", 11)
BUTTON_FONT = ("Segoe UI", 10)

# Top frame with title
header = tk.Label(window, text="📁 Folder to Excel Exporter", font=HEADER_FONT, fg="#003366")
header.pack(pady=(20, 10))

# Description
desc = tk.Label(window,
                text="Select a folder to extract file details and export them to an Excel file.",
                font=LABEL_FONT, wraplength=400, justify="center")
desc.pack(pady=(0, 20))

# Button
browse_button = tk.Button(window, text="Browse Folder", font=BUTTON_FONT, bg="#0066cc", fg="white",
                          padx=20, pady=10, command=browse_folder)
browse_button.pack(pady=(0, 20))

# Status bar
status_var = tk.StringVar()
status_var.set("Ready")
status_bar = tk.Label(window, textvariable=status_var, bd=1, relief=tk.SUNKEN,
                      anchor='w', font=("Segoe UI", 9), bg="#f0f0f0")
status_bar.pack(side=tk.BOTTOM, fill=tk.X)

window.mainloop()
